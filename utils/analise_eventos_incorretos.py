from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
import unicodedata

import pandas as pd
from openpyxl import load_workbook


REQUESTED_RANKINGS: tuple[tuple[str, str], ...] = (
    ("instalador", "Instalador"),
    ("cidade", "Cidade"),
    ("bairro", "Bairro"),
    ("polo", "Polo"),
    ("dia", "Por dia"),
    ("mes", "Mês"),
    ("evento", "Tipo de evento"),
    ("regional", "Regional"),
)


MONTH_NAMES_PT_BR = {
    1: "janeiro",
    2: "fevereiro",
    3: "março",
    4: "abril",
    5: "maio",
    6: "junho",
    7: "julho",
    8: "agosto",
    9: "setembro",
    10: "outubro",
    11: "novembro",
    12: "dezembro",
}


@dataclass(frozen=True)
class EventSheetData:
    dataframe: pd.DataFrame
    header_row: int
    raw_rows: int


def _read_bytes(source: bytes | bytearray | object) -> bytes:
    if isinstance(source, (bytes, bytearray)):
        return bytes(source)
    if hasattr(source, "getvalue"):
        return bytes(source.getvalue())
    if hasattr(source, "read"):
        current_position = source.tell() if hasattr(source, "tell") else None
        data = source.read()
        if current_position is not None and hasattr(source, "seek"):
            source.seek(current_position)
        return bytes(data)
    raise TypeError("Fonte de Excel invalida.")


def _normalize(value: object) -> str:
    text = "" if value is None else str(value)
    text = text.replace("\ufeff", "").strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    return " ".join(text.replace("_", " ").replace("-", " ").split())


def _clean_cell(value: object) -> object:
    if isinstance(value, str):
        text = " ".join(value.replace("\t", " ").split())
        return pd.NA if text == "" else text
    return value


def _unique_headers(values: list[object]) -> list[str]:
    headers: list[str] = []
    used: dict[str, int] = {}
    for index, value in enumerate(values, start=1):
        clean = "" if value is None else str(value).replace("\ufeff", "").strip()
        if not clean:
            clean = f"Coluna {index}"
        count = used.get(clean, 0) + 1
        used[clean] = count
        headers.append(clean if count == 1 else f"{clean}_{count}")
    return headers


def list_excel_sheets(source: bytes | bytearray | object) -> list[str]:
    raw = _read_bytes(source)
    workbook = load_workbook(BytesIO(raw), read_only=True, data_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


HEADER_HINTS = {
    "os",
    "osid",
    "data agendamento",
    "data evento",
    "dtgerada",
    "regional",
    "instalador",
    "evento",
    "observacao",
    "motivo",
    "descricaoevento",
    "lancadapor",
    "cidade",
    "bairro",
    "polo",
    "tipo atibidade os",
    "tipo atividade os",
}


def _header_score(row: tuple[object, ...]) -> tuple[int, int]:
    normalized = [_normalize(value) for value in row if value not in (None, "")]
    non_empty = len(normalized)
    exact_matches = sum(1 for value in normalized if value in HEADER_HINTS)
    partial_matches = sum(
        1
        for value in normalized
        if any(hint in value or value in hint for hint in HEADER_HINTS)
    )
    return exact_matches * 20 + partial_matches * 5 + non_empty, non_empty


def read_excel_sheet(source: bytes | bytearray | object, sheet_name: str) -> EventSheetData:
    raw = _read_bytes(source)
    workbook = load_workbook(BytesIO(raw), read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name]
        scanned_rows = list(
            worksheet.iter_rows(
                min_row=1,
                max_row=min(worksheet.max_row, 30),
                values_only=True,
            )
        )
        if not scanned_rows:
            return EventSheetData(pd.DataFrame(), 1, 0)

        best_index = 0
        best_score = (-1, -1)
        for index, row in enumerate(scanned_rows):
            score = _header_score(row)
            if score > best_score:
                best_index = index
                best_score = score

        header_values = list(scanned_rows[best_index])
        width = 0
        for index, value in enumerate(header_values, start=1):
            if value not in (None, ""):
                width = index
        if width == 0:
            return EventSheetData(pd.DataFrame(), best_index + 1, 0)

        headers = _unique_headers(header_values[:width])
        records: list[list[object]] = []
        blank_streak = 0
        data_started = False
        for row in worksheet.iter_rows(min_row=best_index + 2, values_only=True):
            values = [_clean_cell(value) for value in list(row[:width])]
            is_blank = all(pd.isna(value) for value in values)
            if is_blank:
                blank_streak += 1
                if data_started and blank_streak >= 500:
                    break
                if not data_started and blank_streak >= 1000:
                    break
                continue
            blank_streak = 0
            data_started = True
            records.append(values)

        df = pd.DataFrame(records, columns=headers)
        df = df.dropna(axis=1, how="all")
        return EventSheetData(df, best_index + 1, len(records))
    finally:
        workbook.close()


COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "instalador": (
        "instalador",
        "lancadapor",
        "lancada por",
        "tecnico",
        "responsavel",
        "nome responsavel atendimento",
        "nomeresponsavelatendimento",
        "usuario",
    ),
    "cidade": ("cidade", "city", "municipio"),
    "bairro": ("bairro", "neighborhood"),
    "polo": ("polo", "hub"),
    "regional": ("regional", "regiao"),
    "evento": (
        "evento",
        "motivo",
        "tipo evento",
        "tipo de evento",
        "descricao evento",
        "descricaoevento",
    ),
    "problema": (
        "observacao",
        "obs",
        "motivo bloqueio",
        "motivobloqueio",
        "descricao do bloqueio",
        "descricaodobloqueio",
    ),
    "data": (
        "data evento",
        "data agendamento",
        "dtgerada",
        "data gerada",
        "endtime",
        "data",
    ),
}


def _best_column(columns: list[str], aliases: tuple[str, ...]) -> str | None:
    best_col = None
    best_score = 0
    for col in columns:
        normalized_col = _normalize(col)
        for position, alias in enumerate(aliases):
            normalized_alias = _normalize(alias)
            score = 0
            if normalized_col == normalized_alias:
                score = 100 - position
            elif normalized_alias in normalized_col:
                score = 70 - position
            elif normalized_col in normalized_alias:
                score = 40 - position
            if score > best_score:
                best_score = score
                best_col = col
    return best_col


def detect_event_columns(df: pd.DataFrame) -> dict[str, str | None]:
    columns = [str(col) for col in df.columns]
    return {key: _best_column(columns, aliases) for key, aliases in COLUMN_ALIASES.items()}


def apply_event_date_columns(df: pd.DataFrame, date_col: str | None) -> pd.DataFrame:
    prepared = df.copy()
    if not date_col or date_col not in prepared.columns:
        return prepared
    dates = pd.to_datetime(prepared[date_col], errors="coerce", dayfirst=True)
    if dates.notna().sum() == 0:
        return prepared
    prepared["__evento_data"] = dates.dt.strftime("%d/%m/%Y")
    prepared["__evento_mes"] = dates.apply(
        lambda value: f"{MONTH_NAMES_PT_BR[int(value.month)]}/{int(value.year)}" if pd.notna(value) else pd.NA
    )
    return prepared


def ranking_for_column(df: pd.DataFrame, column: str | None, label: str) -> pd.DataFrame:
    if not column or column not in df.columns:
        return pd.DataFrame(columns=[label, "Quantidade", "% participacao", "% acumulado"])
    values = df[column].astype("string").str.strip()
    values = values[values.notna() & ~values.str.lower().isin(["", "nan", "none", "(tudo)", "(varios itens)", "total geral"])]
    if values.empty:
        return pd.DataFrame(columns=[label, "Quantidade", "% participacao", "% acumulado"])
    ranking = values.value_counts(dropna=True).rename_axis(label).reset_index(name="Quantidade")
    total = float(ranking["Quantidade"].sum())
    ranking["% participacao"] = ranking["Quantidade"] / total * 100
    ranking["% acumulado"] = ranking["% participacao"].cumsum()
    return ranking


def build_event_rankings(df: pd.DataFrame, detected: dict[str, str | None]) -> dict[str, pd.DataFrame]:
    rankings: dict[str, pd.DataFrame] = {}
    prepared = apply_event_date_columns(df, detected.get("data"))
    dimension_columns = {
        "instalador": detected.get("instalador"),
        "cidade": detected.get("cidade"),
        "bairro": detected.get("bairro"),
        "polo": detected.get("polo"),
        "dia": "__evento_data" if "__evento_data" in prepared.columns else None,
        "mes": "__evento_mes" if "__evento_mes" in prepared.columns else None,
        "evento": detected.get("evento"),
        "regional": detected.get("regional"),
    }
    for key, label in REQUESTED_RANKINGS:
        rankings[key] = ranking_for_column(prepared, dimension_columns.get(key), label)
    return rankings


def filter_event_dataframe(
    df: pd.DataFrame,
    detected: dict[str, str | None],
    selected_values: dict[str, list[str]],
    date_range: tuple[pd.Timestamp, pd.Timestamp] | None = None,
) -> pd.DataFrame:
    filtered = df.copy()
    for key, values in selected_values.items():
        column = detected.get(key)
        if not column or column not in filtered.columns or not values:
            continue
        comparable = filtered[column].astype("string").str.strip()
        filtered = filtered[comparable.isin(values)].copy()
    date_col = detected.get("data")
    if date_col and date_col in filtered.columns and date_range:
        dates = pd.to_datetime(filtered[date_col], errors="coerce", dayfirst=True)
        start, end = date_range
        filtered = filtered[(dates >= start) & (dates <= end)].copy()
    return filtered


def format_ranking_table(df: pd.DataFrame) -> pd.DataFrame:
    formatted = df.copy()
    if "Quantidade" in formatted.columns:
        formatted["Quantidade"] = formatted["Quantidade"].map(lambda value: f"{int(value):,}".replace(",", "."))
    for column in ("% participacao", "% acumulado"):
        if column in formatted.columns:
            formatted[column] = formatted[column].map(lambda value: f"{value:.1f}%")
    return formatted


def build_event_excel_bytes(rankings: dict[str, pd.DataFrame], filtered_base: pd.DataFrame) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for key, label in REQUESTED_RANKINGS:
            table = rankings.get(key, pd.DataFrame())
            if not table.empty:
                table.to_excel(writer, sheet_name=label[:31], index=False)
        if not filtered_base.empty:
            filtered_base.to_excel(writer, sheet_name="Base filtrada", index=False)
    return output.getvalue()


def event_insights(ranking: pd.DataFrame, label: str, top_n: int) -> list[str]:
    if ranking.empty:
        return [f"Nao ha dados suficientes para analisar {label.lower()}."]
    top = ranking.iloc[0]
    top_label = str(top[label])
    top_count = int(top["Quantidade"])
    top_pct = float(top["% participacao"])
    visible = ranking.head(top_n)
    accumulated = float(visible["% participacao"].sum())
    concentration = "concentrada" if top_pct >= 35 else "pulverizada"
    count_text = f"{top_count:,}".replace(",", ".")
    return [
        f"{top_label} lidera {label.lower()} com {count_text} eventos ({top_pct:.1f}% dos registros filtrados).",
        f"As {len(visible)} primeiras categorias concentram aproximadamente {accumulated:.1f}% do volume exibido.",
        f"A distribuicao esta {concentration}; avalie o primeiro grupo e tambem a cauda do ranking.",
    ]
